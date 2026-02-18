#!/usr/bin/env python3
import typer
from pathlib import Path
from whoosh.index import create_in, open_dir, exists_in
from whoosh.fields import Schema, TEXT, ID
from whoosh.qparser import QueryParser
from whoosh.highlight import UppercaseFormatter
import PyPDF2
from docx import Document
from openpyxl import load_workbook

app = typer.Typer(help="本地文档智能搜索工具")

# 定义索引结构
schema = Schema(
    path=ID(stored=True, unique=True),
    content=TEXT(stored=True)
)

def extract_pdf(file_path):
    """提取PDF文本"""
    try:
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            return ' '.join(page.extract_text() for page in reader.pages)
    except:
        return ""

def extract_docx(file_path):
    """提取DOCX文本"""
    try:
        doc = Document(file_path)
        return ' '.join(p.text for p in doc.paragraphs)
    except:
        return ""

def extract_xlsx(file_path):
    """提取XLSX文本"""
    try:
        wb = load_workbook(file_path, data_only=True)
        text = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                text.extend(str(cell) for cell in row if cell)
        return ' '.join(text)
    except:
        return ""

@app.command()
def index(
    directory: Path = typer.Argument(..., help="要扫描的目录"),
    index_dir: Path = typer.Option(".docfind_index", help="索引存储目录")
):
    """扫描目录并建立文档索引"""
    if not directory.exists():
        typer.echo(f"❌ 目录不存在: {directory}", err=True)
        raise typer.Exit(1)
    
    index_dir.mkdir(exist_ok=True)
    
    # 创建或打开索引
    if not exists_in(str(index_dir)):
        ix = create_in(str(index_dir), schema)
    else:
        ix = open_dir(str(index_dir))
    
    writer = ix.writer()
    count = 0
    
    # 扫描文档
    extensions = {'.pdf': extract_pdf, '.docx': extract_docx, '.xlsx': extract_xlsx}
    
    with typer.progressbar(
        list(directory.rglob('*')),
        label="扫描文档"
    ) as files:
        for file_path in files:
            if file_path.suffix.lower() in extensions:
                typer.echo(f"索引: {file_path.name}")
                content = extensions[file_path.suffix.lower()](file_path)
                if content.strip():
                    writer.add_document(
                        path=str(file_path.absolute()),
                        content=content
                    )
                    count += 1
    
    writer.commit()
    typer.echo(f"✅ 完成！已索引 {count} 个文档")

@app.command()
def search(
    query: str = typer.Argument(..., help="搜索关键词"),
    index_dir: Path = typer.Option(".docfind_index", help="索引目录"),
    limit: int = typer.Option(10, help="最多显示结果数")
):
    """搜索文档内容"""
    if not exists_in(str(index_dir)):
        typer.echo("❌ 索引不存在，请先运行 'docfind index' 建立索引", err=True)
        raise typer.Exit(1)
    
    ix = open_dir(str(index_dir))
    
    with ix.searcher() as searcher:
        query_obj = QueryParser("content", ix.schema).parse(query)
        results = searcher.search(query_obj, limit=limit)
        results.formatter = UppercaseFormatter()
        
        if not results:
            typer.echo("😕 没有找到匹配的文档")
            return
        
        typer.echo(f"\n🔍 找到 {len(results)} 个结果:\n")
        
        for i, hit in enumerate(results, 1):
            typer.echo(f"[{i}] {Path(hit['path']).name}")
            typer.echo(f"    路径: {hit['path']}")
            typer.echo(f"    相关度: {hit.score:.2f}")
            
            # 显示匹配片段（高亮）
            highlights = hit.highlights("content", top=3)
            if highlights:
                typer.echo(f"    片段: {highlights[:200]}...")
            typer.echo()

if __name__ == "__main__":
    app()